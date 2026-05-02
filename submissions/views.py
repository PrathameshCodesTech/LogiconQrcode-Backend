from datetime import datetime
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from rest_framework import generics, permissions, status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.throttling import AnonRateThrottle

from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Submission
from .serializers import (
    SubmissionCreateSerializer, SubmissionListSerializer,
    SubmissionDetailSerializer, SubmissionStatusUpdateSerializer,
)
from .services import (
    create_submission_documents,
    create_submission_with_answers,
    create_review_log,
)
from .filters import SubmissionFilter


class PublicSubmissionThrottle(AnonRateThrottle):
    scope = 'public_submission'


class PublicSubmissionCreateView(APIView):
    permission_classes = [permissions.AllowAny]
    throttle_classes = [PublicSubmissionThrottle]

    def post(self, request):
        if 'resume' not in request.FILES:
            return Response(
                {'resume': ['Resume is required.']},
                status=status.HTTP_400_BAD_REQUEST
            )
        serializer = SubmissionCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        with transaction.atomic():
            submission = create_submission_with_answers(serializer.validated_data, request)
            documents = create_submission_documents(submission, request.FILES)
        return Response(
            {
                'id': submission.id,
                'status': submission.status,
                'is_possible_duplicate': submission.is_possible_duplicate,
                'documents_count': len(documents),
                'message': 'Submission received successfully.',
            },
            status=status.HTTP_201_CREATED,
        )


class AdminSubmissionListView(generics.ListAPIView):
    queryset = Submission.objects.select_related('campaign', 'role', 'site', 'candidate').all()
    serializer_class = SubmissionListSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = SubmissionFilter
    search_fields = ['full_name', 'mobile_number', 'mobile_number_normalized']
    ordering_fields = ['submitted_at', 'status', 'full_name']
    ordering = ['-submitted_at']

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params

        search = params.get('search', '').strip()
        if search:
            qs = qs.filter(
                Q(full_name__icontains=search) |
                Q(mobile_number__icontains=search) |
                Q(mobile_number_normalized__icontains=search)
            )

        status_val = params.get('status', '')
        if status_val:
            qs = qs.filter(status=status_val)

        date_from = params.get('date_from', '')
        if date_from:
            qs = qs.filter(submitted_at__gte=date_from)

        date_to = params.get('date_to', '')
        if date_to:
            qs = qs.filter(submitted_at__lte=date_to)

        role_id = params.get('role_id', '')
        role_filter = params.get('role_filter', '')

        if role_filter == 'other':
            qs = qs.filter(role__isnull=True).exclude(other_role_title_normalized='')
        elif role_id:
            qs = qs.filter(role_id=role_id)

        return qs


class AdminSubmissionDetailView(generics.RetrieveAPIView):
    queryset = Submission.objects.select_related(
        'campaign', 'role', 'site', 'candidate'
    ).prefetch_related(
        'answers__field', 'documents', 'reviews__reviewed_by'
    ).all()
    serializer_class = SubmissionDetailSerializer
    permission_classes = [permissions.IsAuthenticated]


class AdminSubmissionStatusUpdateView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, pk):
        try:
            submission = Submission.objects.get(pk=pk)
        except Submission.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = SubmissionStatusUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        create_review_log(
            submission,
            request.user,
            serializer.validated_data['status'],
            serializer.validated_data.get('note', ''),
        )
        return Response({'id': submission.id, 'status': submission.status})


class AdminSubmissionResumeUrlView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        try:
            submission = Submission.objects.get(pk=pk)
        except Submission.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        resume_doc = submission.documents.filter(document_type='resume').first()
        if not resume_doc or not resume_doc.file:
            return Response({'resume_url': None})

        return Response({'resume_url': resume_doc.file.url})


class AdminSubmissionExportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def _apply_filters(self, qs, params):
        search = params.get('search', '').strip()
        if search:
            qs = qs.filter(
                Q(full_name__icontains=search) |
                Q(mobile_number__icontains=search) |
                Q(mobile_number_normalized__icontains=search)
            )

        status_val = params.get('status', '')
        if status_val:
            qs = qs.filter(status=status_val)

        date_from = params.get('date_from', '')
        if date_from:
            qs = qs.filter(submitted_at__gte=date_from)

        date_to = params.get('date_to', '')
        if date_to:
            qs = qs.filter(submitted_at__lte=date_to)

        campaign_id = params.get('campaign_id', '')
        if campaign_id:
            qs = qs.filter(campaign_id=campaign_id)

        site_id = params.get('site_id', '')
        if site_id:
            qs = qs.filter(site_id=site_id)

        role_id = params.get('role_id', '')
        role_filter = params.get('role_filter', '')

        if role_filter == 'other':
            qs = qs.filter(role__isnull=True).exclude(
                other_role_title_normalized=''
            )
        elif role_id:
            qs = qs.filter(role_id=role_id)

        return qs

    def _get_applied_role_display(self, sub):
        if sub.role:
            return sub.role.name
        if sub.other_role_title:
            return f"Other: {sub.other_role_title}"
        return ''

    def _get_role_count(self, sub):
        same_mobile_subs = Submission.objects.filter(
            campaign=sub.campaign,
            mobile_number_normalized=sub.mobile_number_normalized,
        )
        role_ids = set()
        other_titles = set()
        for s in same_mobile_subs:
            if s.role_id:
                role_ids.add(s.role_id)
            elif s.other_role_title_normalized:
                other_titles.add(s.other_role_title_normalized)
        return len(role_ids) + len(other_titles)

    def _build_dynamic_columns(self, qs):
        raw_columns = []
        seen_keys = set()

        for sub in qs.prefetch_related('answers'):
            for ans in sub.answers.all():
                key = ans.field_id or ans.field_label_snapshot
                if not key or key in seen_keys:
                    continue
                seen_keys.add(key)
                raw_columns.append({
                    'key': key,
                    'base_label': ans.field_label_snapshot or f"Field {ans.field_id}",
                    'header': ans.field_label_snapshot or f"Field {ans.field_id}",
                })

        label_seen = {}
        for column in raw_columns:
            base_label = column['base_label']
            count = label_seen.get(base_label, 0)
            if count:
                column['header'] = f"{base_label} ({count + 1})"
            label_seen[base_label] = count + 1

        return raw_columns

    def get(self, request):
        params = request.query_params

        qs = Submission.objects.select_related(
            'campaign', 'role', 'site', 'candidate'
        ).prefetch_related('answers', 'documents').all()

        qs = self._apply_filters(qs, params)

        fixed_columns = [
            'Submission ID', 'Submitted At', 'Status', 'Language',
            'First Name', 'Middle Name', 'Last Name', 'Full Name',
            'Mobile Number', 'Campaign', 'Site', 'Applied Role',
            'Other Role Title', 'Possible Duplicate', 'Duplicate Reason',
            'No. of Roles', 'Resume Uploaded', 'Resume URL', 'Documents Count',
        ]

        dynamic_columns = self._build_dynamic_columns(qs)
        all_columns = fixed_columns + [col['header'] for col in dynamic_columns]

        wb = Workbook()
        ws = wb.active
        ws.title = 'Submissions'

        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="D9E1F2")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col_idx, col_name in enumerate(all_columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        for row_idx, sub in enumerate(qs, start=2):
            sub_answers = {
                (ans.field_id or ans.field_label_snapshot): ans.value
                for ans in sub.answers.all()
            }

            resume_doc = sub.documents.filter(document_type='resume').first()
            resume_url = ''
            if resume_doc and resume_doc.file:
                resume_url = request.build_absolute_uri(resume_doc.file.url)

            row_data = [
                sub.id,
                sub.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
                sub.status,
                sub.language,
                sub.first_name,
                sub.middle_name,
                sub.last_name,
                sub.full_name,
                sub.mobile_number,
                sub.campaign.title if sub.campaign else '',
                sub.site.name if sub.site else '',
                self._get_applied_role_display(sub),
                sub.other_role_title,
                'Yes' if sub.is_possible_duplicate else 'No',
                sub.duplicate_reason or '',
                self._get_role_count(sub),
                'Yes' if resume_doc else 'No',
                resume_url,
                sub.documents.count(),
            ]

            fixed_column_count = len(fixed_columns)
            for col_idx, col_name in enumerate(all_columns, start=1):
                value = row_data[col_idx - 1] if col_idx <= len(row_data) else ''

                if col_idx > fixed_column_count:
                    dynamic_col = dynamic_columns[col_idx - fixed_column_count - 1]
                    val = sub_answers.get(dynamic_col['key'])
                    if isinstance(val, list):
                        value = ', '.join(str(v) for v in val)
                    else:
                        value = str(val) if val is not None else ''

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        for col_idx in range(1, len(all_columns) + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=min(100, ws.max_row)):
                cell = row[0]
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

        ws.freeze_panes = 'A2'

        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="logicon_submissions_{timestamp}.xlsx"'
        wb.save(response)
        return response
